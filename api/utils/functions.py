import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import pandas as pd
import math

def generate_quarter_code(input_date,day_of_month):

    if 1 <= day_of_month <= 15:
        quarter_code = '1QNZ'
    elif 16 <= day_of_month <= 31:
        quarter_code = '2QNZ'
    else: 
        quarter_code = 'ERROR DATE'
    # French month names
    french_month_names = [
    'janv', 'févr', 'mars', 'avr', 'mai', 'juin',
    'juil', 'août', 'sept', 'oct', 'nov', 'déc']
    # Extract month abbreviation in French and last two digits of the year
    month_abbreviation = french_month_names[input_date.month - 1]
    year_last_digits = input_date.strftime("%y")
    # Combine the results
    formatted_date = f"{month_abbreviation} {year_last_digits}"
    # Concatenate quarter_code and formatted_date
    result = quarter_code + " " + formatted_date

    return result

# find the code aux for a specific client 
def get_code_aux(client,aux_file):

    # Client lowercase and Normalized
    # the LIBELLE black column normalized
    aux_file_nor = aux_file['LIBELLE black'].apply(lambda x: ' '.join(str(x).split())).str.lower()
    client = client.lower()
    client = ' '.join(client.split())

    # aux_file lowercase and normalized 
    if client in aux_file_nor.values:
        code = int (aux_file.loc[(aux_file_nor == client),'CODE'].iloc[0])
        return str(code)
    else:

        # Client not found, return empty string
        return ""
    # get the sum of COM. Str. for one client 
def cal_sum_com_str (client, billet_file):
    

    # Filter the DataFrame to include only rows where the 'Client' column is equal to the target client
    filtered_df = billet_file[billet_file['Client'] == client]

    # Calculate the sum of the 'Com. Str.' column for the filtered DataFrame
    sum_com_str = filtered_df['Com. Str.'].sum() 

    return sum_com_str

def cal_sum_montant_billet(client, billet_file):
    # Filter the DataFrame to include only rows where the 'Client' column is equal to the target client
    filtered_df = billet_file[billet_file['Client'] == client]
     # Calculate the sum of the 'Com. Str.' column for the filtered DataFrame
    sum_montant = filtered_df['Montant'].sum() 

    return sum_montant

def cal_sum_facture(client, fact_file):
    # Filter the DataFrame to include only rows where the 'Client' column is equal to the target client
    filtered_df = fact_file[fact_file['Clients'] == client]
     # Calculate the sum of the 'facture.' column for the filtered DataFrame
    sum_facture = filtered_df['Facture'].sum() 

    sum_avoir = filtered_df['Avoir'].sum() 

    sum = sum_facture - sum_avoir


    return sum

def get_code_com(client,com_file):
    # Normalize Client and LIBELLE column in com_file 
    com_file_nor = com_file['LIBELLE'].apply(lambda x: ' '.join(str(x).split())).str.lower()
    client = client.lower()
    client = ' '.join(client.split())

    if client in com_file_nor.values:

        code = int (com_file.loc[(com_file_nor== client.lower()),'CODE'].iloc[0])
        return str(code)
    else:
        # Client not found, return empty string
        return ""
    
def comapre_billet_facture(file_billet, file_facture):
    # Extract the unique values from columnA in file
    values_in_billet = set(file_billet['Client'])

   # Extract the unique values from columnB in fileB
    values_in_facture = set(file_facture['Clients'])

   # Find values in columnA that are not in columnB
    values_only_in_billet = values_in_billet - values_in_facture

    print('listes des billets ')
    for value in values_only_in_billet:
        print(f"{value}\n")

    return values_only_in_billet

# calculate the diffrence between sum montant billet and montant facture 
 
def calcule_diff (sum_montant_billet, sum_fact) :
     diff  = sum_montant_billet - sum_fact
     if abs(diff) < 0.1:
         diff = 0

     return diff          

def create_journal_17(billet_file, fact_file,aux_file, com_file):
        
    # create journal 17 workbook 
    journal17_workbook = openpyxl.Workbook()
    new_sheet = journal17_workbook.active
    labels = ['FOLIO',	'PIECE', 'Date', 'REFERENCE',	'LIBELLE',	'CODE_JRN',	'CODE_COM',	'CODE_AUX',	'DEBIT',	'CREDIT']
    # Write labels to the first row
    new_sheet.append(labels)

    # create a new file for rapprochement  
    rapprochement_workbook = openpyxl.Workbook()
    new_sheet_raps = rapprochement_workbook.active
    # first row of journal 17 file 
    labels_rapprochement = ['Client',	'Montant billets', 'Montant facture', 'Différence']
    # Write labels to the first row
    new_sheet_raps.append(labels_rapprochement)
    

    fact_columns = [ 'N° Facture','Clients', 'Facture','Avoir'] 
    # dropna() method will delete all the rows that has NAN values for the fact_columns 
    # fact_columns and avoir 
    fact_file_cleaned = fact_file.dropna(subset=fact_columns, how='all')

    #print(fact_file_cleaned[fact_columns])
    # get one value of Date facture column 

    first_date_value = fact_file_cleaned['Date Facture'].iloc[0]
    print (first_date_value )

    # Convert to datetime object
    input_date = datetime.strptime(first_date_value, "%Y-%m-%d")

    # Extract the month number
    month_number = input_date.month

    # Print the result
    print(month_number)

    # Convert to a string and pad with zeros
    formatted_month = str(month_number).zfill(5)

    # Print the result
    print(formatted_month)

    # Replace hyphens with forward slashes
    formatted_date = first_date_value.replace('-', '/')

    # Print the result
    print(formatted_date)

    # Check the day of the month
    day_of_month = input_date.day

    # Print the result
    reference = generate_quarter_code(input_date,day_of_month)
    print(reference)


    cstat = 'CSTAT FACTURE'

    # all the clients in facture file 
    list_clients = []
    i = 0
    for index, fact_row in fact_file_cleaned.iterrows():
     # index is the index label of the current row being iterated over the Dataframe 
     i = i+1 
     # Uncomment the line below if you want to print the entire row
     #print(fact_row)
     num_facture = fact_row['N° Facture']
     client = fact_row['Clients'] 
     #add the client to the array list of clients 
     list_clients.append(client)

     debit = fact_row['Facture']
     avoir = fact_row['Avoir']
     print(f"N° Facture: {num_facture}, Client: {client}, facture: {debit}")

     cstat_client_facture = cstat + " " + client + "  " + num_facture
     #calculate the value of Com. Str . for this client  and mutiply by -1 
     credit = (cal_sum_com_str(client, billet_file)) * -1
    
  
     code_aux = get_code_aux(client,aux_file)

     if code_aux != "":
        # Client exists in TAB_AUX 
        code_com = '411000'
     else: 
        # code_com = code , or code_com = null 
        code_com = get_code_com(client, com_file)
        #code_com = "error"

     # Rembourssement / Avoir 
     if (avoir != 0) and not (math.isnan(avoir)):
        debit = avoir * -1
        print(f"Avoir de {debit} pour {client}") 

     # if code is not found you need to search in another file, that file is tab_com 
     # in this case the code_aux is null empty " " but the first code is the one in tab_com 
     data_credit= [month_number,	formatted_month, formatted_date,	reference,	cstat_client_facture,	'17'	,'701001',	code_aux ,'0', credit]
     data_debit= [month_number,	formatted_month, formatted_date,	reference,	cstat_client_facture,	'17'	,code_com,	code_aux , debit, '0']

     new_sheet.append(data_credit)
     new_sheet.append(data_debit)

     # comparer la diference entre facture et montant pour chaque client 
     sum_montant_billet = cal_sum_montant_billet(client, billet_file)
     sum_fact = cal_sum_facture(client, fact_file)
     diff  = calcule_diff (sum_montant_billet, sum_fact)
     if diff != 0:
        data_rap = [client , sum_montant_billet , sum_fact, diff]
        new_sheet_raps.append(data_rap)


    # find all clients that only exist in liste des billets 

    client_billet = comapre_billet_facture(billet_file, fact_file)

    for client in client_billet:
        sum_montant_billet = cal_sum_montant_billet(client, billet_file)
        sum_fact = 0 
        diff = calcule_diff (sum_montant_billet, sum_fact)

    if sum_montant_billet != 0:
        data_billet = [client , sum_montant_billet , sum_fact, diff]
        new_sheet_raps.append(data_billet)

    return journal17_workbook,rapprochement_workbook     


__all__ = ['generate_quarter_code', 'get_code_aux','cal_sum_com_str','cal_sum_montant_billet','cal_sum_facture', 'get_code_com','comapre_billet_facture', 'calcule_diff', 'create_journal_17'  ]